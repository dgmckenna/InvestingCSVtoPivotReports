#! python3

# Updated 2016-11
# Works with the "new" webbroker csv export file format.


# Ribbon_CSV Personal Finance Python Scripting Program
# by David McKenna
#
#
# This script walks through the current working directory
# looking for all csv files. It will take TD Waterhouse
# downloaded CSV files, and extract the date, account information
# and list of securities from each file. The script consolidates
# all of the information together, and writes it into an
# excel file. The goal of this is all of the data is in one file
# with columns for date, account and a vlookup to create a column
# for "category" (ie. US equity, fixed income). This allows the use of
# pivot tables and pivot charts to easily look at such things as:
# current asset allocation, asset allocation by account (how much
# international equity in my RRSP, etc.
#
# To support financial information that may be not available
# in TD waterhouse csv format, the excel file has an "offline" tab.
# any entries made by hand into the "offline" tab are copied into
# the source data tab each time the script is run. This is useful
# if you have GIC's, assets at another institution or other items
# you wish to track.
#
# Note: you cannot have the pivot table, in the excel file that
# this script builds. This script uses a python library which
# cannot handle pivot tables. Instead, you have to have a separate
# file with a pivot table, but it reads it's source data
# from the file built by this Python script.

import pprint
import os
import sys
import datetime
import re
import string

import csv
import openpyxl

from openpyxl.cell import get_column_letter

# Constants for excel file
SOURCE_DATA_TAB = 'SourceData'
ASSET_ALLOC_TAB = 'AssetAllocationLookup'
OFFLINE_DATA_TAB = 'Offline'
PIVOT_DATA_NAME_RANGE = 'PivotData'
ASSET_ALLOC_NAME_RANGE = 'AssetAllocation'


def is_number(s):
# Routine to check if some values from the CSV file are numeric or not
# If they are numeric we need to format them differently when written
# to an excel file.

    try:
        float(s)
        return True
    except ValueError:
        return False


def setup_excel_file(output_excel_filename):
# This routine creates the initially required excel file if it doesn't
# exist. The routine creates the excel file that holds all of the
# pivot table data, but does not create the excel file that holds
# the pivot table itself. Currently that is not possible with the
# openpyxl library.

    if not(os.path.exists(output_excel_filename)):
        #Create workbook with the required tabs
        wb = openpyxl.Workbook()
        wb.get_active_sheet().title = SOURCE_DATA_TAB
        wb.create_sheet (index=1, title=ASSET_ALLOC_TAB)
        wb.create_sheet (index=3, title=OFFLINE_DATA_TAB)
        wb.save (output_excel_filename)


def initialize_pvt_data(pvt_data):
# This routine sets the columns that will appear in the pivot
# table data.
    

    
    header_row = [
        'Date', 'Account' , 'Account Currency', 'Account Type',
        'Symbol', 'Security',  'Quantity',  'Price',  'Book Value',
        'Market Value',  'Unrealized $',  'Gain/Loss %'
        ]
    header_row.insert(7,'Category')
    pvt_data.append(header_row)


def read_CSV_file(CSVFilePath, security_dict, pvt_data):
# This routine opens a CSV file (that has already been identified as
# a TD waterhouse file) and extracts the date, account information
# and security information. The extracted information is 
# appended to the pvt_data list.
    
    csv_reader_file = open(CSVFilePath)
    csv_reader = csv.reader(csv_reader_file)
    csv_data = list (csv_reader)
    
    # The CSV file structure is:
    # [As of Date,2016-10-21 11:17:48],
    # [Account,TD Direct Investing - 538R77A],   (Note: CDN Cash might be TFSA, or RRSP)
    # ['Cash Balance (after settlement)', '[numeric_value]', '', '', '', ''],
    # ['Securities Market Value', '[numeric_value]', '', '', '', ''],
    # ['Total Account Value', '[numeric_value]', '', '', '', ''],
    # ['Margin Available (as of yesterday)', 'N/A', '', '', '', ''],
    # [],
    # ['Symbol', 'Security',  'Quantity',  'Price',  'Book Value',  'Market Value',  'Unrealized $',  'Gain/Loss %',  '% of Holdings'],
    # <Security data rows follow>

    # Extract the date from the file, and re-format it to a date time object
    # in YYYY-MM-DD
    
    file_date_string = csv_data[0][1].split()[0]
    file_date = datetime.datetime.strptime(file_date_string, '%Y-%m-%d').strftime('%Y-%m-%d')

    # Extract the account number information and split it into a list
    # Account info list:
    # 0=account number, 2=Currency, 3=account type SDRSP/TFSA/ETC. XXXXXXX - CDN TFSA TD Direct Investing

    account_string = csv_data[1][1]
    account_info_list = account_string.split(' ') 
    
    # Loop through all of the securities information in the CSV file.
    # We skip the first 8 rows to get directly to the securities information.
    # For each item, build a record to insert into our pvt data, by taking
    # the date and account information (already extracted from the start
    # of the file) and the rest of the security information from the row.
    for i in range(8, len(csv_data)):
        # Delete the % holding info, which was copied from the csv file. We do not use this
        # item from the export file, and it simplifies building the row data.
        del csv_data[i][len(csv_data[i])-1]

        row = []
        row.append (file_date) #Date
        row.append (account_info_list[4]) #account number
        row.append ('') #currency
        if account_info_list[4][-1] == 'S':
            row.append('SDRSP')
        elif account_info_list[4][-1] == 'J':
            row.append('TFSA')
        else:
            row.append('Cash')
        
        #append for account type
        

        # Copy all items from the exported CSV file, (except %holding which deleted),
        # into our row. This provides data like name, description, quantity, etc.
        row.append(csv_data[i][0]) #symbol
        row.append(csv_data[i][2]) #desc
        row.append(csv_data[i][3]) #qty
        row.append(csv_data[i][5]) #price
        #row.append(csv_data[i][4]) #cost
        row.append(csv_data[i][6]) #book value
        row.append(csv_data[i][7]) #market value
        row.append(csv_data[i][8]) #unrealized $
        row.append(csv_data[i][9]) #unrealized percentage
        
        row.insert(7,'') #placeholder for the Category, which comes from a Vlookup added later
        pvt_data.append (row)
        print (row)

        # Extract the security description, and add it to a dictionary.
        # We need to build a list of all the securities that are in the files
        # so we can generate the asset allocation tab, where there is one
        # row, per security and it's category is listed (ie. US equity, Canadian equity, etc.)
        # The description is used instead of the symbol, to account for the export
        # files not having the market (ie. Canadian Market, or US market, or other). There
        # may be issues where two different securities would have the same symbol on different
        # markets, however it is highly unlikely they would have the same description.
        
        security_desc = csv_data[i][2]  #changed from [1]
        security_symbol = csv_data[i][0]
        
        #make a dictionary with description as the key, then symbol as the value
        security_dict[security_desc] = security_symbol




def write_source_data_tab(wb, pvt_data):
# This routine takes the pivot table data in a list and writes it to the
# excel file. It also copies any data from the "offline" tab and puts it
# in the source data tab. The source data tab is completely regenerated
# each time it runs, so do not delete csv files after this has ran.

    wb.remove_sheet (wb.get_sheet_by_name(SOURCE_DATA_TAB))
    wb.create_sheet (index=0, title=SOURCE_DATA_TAB)
    sheet = wb.get_sheet_by_name(SOURCE_DATA_TAB)
    letters = list(string.ascii_lowercase)

    #From the data extracted from the CSV, populate the tab with
    #data that could be used in a pivot table format
    for security_index in range (0, len(pvt_data)):
        print (pvt_data[security_index])
        for col in range (0, len(pvt_data[0])):
            #print (col)
            if col == 0 and security_index > 0:
                date_info_temp = pvt_data[security_index][col].split('-')
                sheet[letters[col]+str(security_index+1)] = datetime.datetime(
                                                                                int(date_info_temp[0]),
                                                                                int(date_info_temp[1]),
                                                                                int(date_info_temp[2])
                                                                                )
            
            elif is_number(pvt_data[security_index][col]):
                #Added float checking and casting to have numbers as numbers instead of text
                sheet[letters[col]+str(security_index+1)] = (
                    float(pvt_data[security_index][col])
                    )
            else:
                sheet[letters[col]+str(security_index+1)] = (
                    pvt_data[security_index][col]
                    )
            if col == 7 and security_index > 0: #If we are not on the header row, then setup the vlookup to get the asset allocation
                sheet[letters[col]+str(security_index+1)].value = (
                    '=vlookup(F' + str(1+security_index) + ',' + ASSET_ALLOC_NAME_RANGE + ',2,false)'
                    )

def write_asset_alloc_tab(wb, security_dict):


    existing_securities_list = []
    new_securities = []
    # Rewrite the tab to have the correct headers
    # load all securities from the existing asset allocation tab
    # compare the securities already in the tab, with the securities in the csv
    # identify new securities, and write them into the tab with "undefined"
    # as the asset allocation
    
    #If you cannot find it, then add it in, and mark it as "undefined"
    sheet = wb.get_sheet_by_name(ASSET_ALLOC_TAB) 
    sheet['A1'] = 'Symbol'
    sheet['B1'] = 'Security'
    sheet['C1'] = 'Category'

    #Build a list of existing securities in the asset allocation tab

    if sheet.get_highest_row() > 1:
        existing_securities = tuple(sheet['B2':'B' + str(sheet.get_highest_row())])
        for row_of_cell_objects in existing_securities:
            for cell in row_of_cell_objects:
                existing_securities_list.append(cell.value)
    else:
        existing_securities_list = []

    #Compare the list of existing securities against the list of securities in the security_dict
    #security dict it loaded from the sourceData, and from the offline tab

    print (security_dict)
    
    for security in security_dict.keys():
        if security not in existing_securities_list:
            new_securities.append(security)

    int_insert_row = sheet.get_highest_row() + 1

    # Write the new securities into the list of securities tab on the excel file
    # mark them as category "undefined"
    for security in new_securities:
        sheet['B' + str(int_insert_row)] = security
        sheet['A' + str(int_insert_row)] = security_dict[security]
        sheet['C' + str(int_insert_row)] = 'undefined'
        int_insert_row = int_insert_row + 1

    if len(new_securities) > 0:
        print ('New securities since file ran!!!')
        pprint.pprint(new_securities)


def copy_offline_data(wb, security_dict):
# This routine copies data from the offline data tab into the
# main tab with all of the security information from the
# waterhouse CSV files. This is done in case you have securities
# or investments at another institution, and have to manually
# lookup the information for that institution and type it into
# this sheet, but still want to combine it together for reporting.

    offline_sheet = wb.get_sheet_by_name(OFFLINE_DATA_TAB)
    source_sheet = wb.get_sheet_by_name(SOURCE_DATA_TAB)
    
    offline_sheet['A1'] = 'Date'
    offline_sheet['B1'] = 'Account'
    offline_sheet['E1'] = 'Symbol'
    offline_sheet['F1'] = 'Description(Used for lookup)'
    offline_sheet['J1'] = 'Book Value'
    offline_sheet['K1'] = 'Market Value'

    int_insert_row = int(source_sheet.get_highest_row())
    int_insert_col = 1

    # Load all the Symbol(Vlookup) names from the offline into the security_dict, this ensures they
    # are added to the AssetAllocationLookup with "undefined" if necessary
    
    if offline_sheet.get_highest_row() > 1:
        offline_securities = tuple(offline_sheet['F2':'F' + str(offline_sheet.get_highest_row())])
        for row_securities in offline_securities:
            for cell in row_securities:
                security_dict[cell.value] = cell.value
        

    # If the offline sheet has items, loop through each row and copy the values into the
    # source data sheet (where we build the pivot table source data).
    if offline_sheet.get_highest_row() > 1:
        offline_securities = tuple(offline_sheet['A2':'K' + str(offline_sheet.get_highest_row())])
        for row in offline_securities:
            int_insert_col = 1
            int_insert_row += 1
            for cell in row:
                source_sheet.cell(row=int_insert_row, column=int_insert_col).value = cell.value
                int_insert_col += 1

            #Add logic so we setup the vlookup properly
            source_sheet.cell(row=int_insert_row, column=8).value = (
                '=vlookup(F' + str(int_insert_row) + ',' + ASSET_ALLOC_NAME_RANGE  + ',2,false)'
                )
                
    

def setup_named_ranges(wb):
    # Setup a named range called "PivotData"
    # When using an excel file, with a pivot table that points at the source data file
    # It's useful to have a named range so you never have to reset your pivot
    # table data range, when you have more data.

    source_sheet = wb.get_sheet_by_name(SOURCE_DATA_TAB)
    wb.create_named_range (
        PIVOT_DATA_NAME_RANGE,source_sheet,
        'A1:' + get_column_letter(source_sheet.get_highest_column())
            + str(source_sheet.get_highest_row())
        )

    # This second named range, is created for the vlookup in the source data.
    # The source data does a vlookup of the asset allocation information
    
    asset_sheet = wb.get_sheet_by_name(ASSET_ALLOC_TAB)
    wb.create_named_range (
        ASSET_ALLOC_NAME_RANGE, asset_sheet,
        'B2:' + get_column_letter(asset_sheet.get_highest_column())
            + str(asset_sheet.get_highest_row())
        )
      




# Start of main program



source_path = ''
output_excel_filename = 'PythonExcel.xlsx'

print ('Start of program!!!')
print ('***')
print (source_path)
print ('***')


# Create the initial excel file if not present
setup_excel_file(output_excel_filename)
pvt_data_list = []
initialize_pvt_data(pvt_data_list)
security_dict = {}
found_one_file = False

for foldername, subfolders, filenames in os.walk(os.getcwd()):
    for filename in filenames:
        #search for CSV export files from waterhouse
        #file must be numbers (typically 3, but not always), followed by a letter, followed by 2 numbers, followed by A,S,J or B
        #export CSV files from waterhouse are named (account number)-30-Jun-2015.csv
        matchWaterhouseFile = re.search(r'^\d+\w\d+[ASJB]', filename)
        if filename.endswith('.csv'):
            print ('Processing ' + os.path.join(foldername, filename))
            read_CSV_file (os.path.join(foldername, filename), security_dict, pvt_data_list)
            found_one_file = True

wb = openpyxl.load_workbook(output_excel_filename)
# After parsing all of the csv files write the source data tab with
# all the csv file data
write_source_data_tab(wb, pvt_data_list)

# Copy any offline data from the offline data tab into the source
# data tab.
copy_offline_data(wb, security_dict)

# Update/generate the asset allocation tab, which lists each
# security once and is used with a vlookup to identify the category.
write_asset_alloc_tab(wb,security_dict)

# Update the named ranges
setup_named_ranges(wb)
wb.save (output_excel_filename) 

if found_one_file == False:
    print(
        'No Waterhouse CSV files found.\n'
        'Put Waterhouse CSV files in this folder, or subfolders of this folder.\n'
        )
           

print('Press Enter to Continue')    
input()
